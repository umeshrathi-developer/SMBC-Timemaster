"""Utility functions for timesheet and holiday import"""
from openpyxl import load_workbook
from datetime import datetime, date
from decimal import Decimal
from django.core.exceptions import ValidationError
import logging
import re
from .models import TimesheetSummary, TimesheetDetails, AttendanceSummary, AttendanceDetails, Holiday, TimesheetEntry, Employee

import_logger = logging.getLogger('timesheet.import')


def _normalize_name(value):
    """Normalize Excel/header names for reliable employee matching."""
    return re.sub(r'\s+', ' ', str(value or '').replace('\xa0', ' ')).strip()


def _employee_name_from_header(value):
    """Extract employee name from headers like 'Employee Name\\nClient Id'."""
    lines = [
        _normalize_name(line)
        for line in str(value or '').replace('\xa0', ' ').splitlines()
        if _normalize_name(line)
    ]
    if lines:
        return lines[0]
    return _normalize_name(value)


def _column_letter(column_index):
    """Return Excel-style column letters for a zero-based column index."""
    column_index += 1
    letters = ''
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _record_client_import_issue(results, bucket, **details):
    results[bucket].append({
        'row': details.get('row', ''),
        'column': details.get('column', ''),
        'date': details.get('date', ''),
        'employee': details.get('employee', ''),
        'raw_header': details.get('raw_header', ''),
        'value': details.get('value', ''),
        'reason': details.get('reason', ''),
        'project': details.get('project', ''),
    })


def import_holiday_file(file_obj):
    """
    Import holiday data from an Excel file.

    Expected columns:
    - S. No.
    - Date
    - Day
    - Holiday
    - Applicability

    Column mapping:
    - Holiday -> Holiday.name
    - Applicability -> Holiday.location
    - holiday_type -> PUBLIC_HOLIDAY
    """
    try:
        workbook = load_workbook(file_obj)
        worksheet = workbook.active
        results = {
            'success': True,
            'message': 'Holiday import completed successfully',
            'created_count': 0,
            'updated_count': 0,
            'errors': [],
        }

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {
                'success': False,
                'message': 'The uploaded file is empty.',
                'errors': ['Missing header row.'],
            }

        header_map = {
            str(column).strip().lower(): index
            for index, column in enumerate(header_row)
            if column is not None
        }
        required_headers = ['date', 'holiday', 'applicability']
        missing_headers = [header for header in required_headers if header not in header_map]
        if missing_headers:
            return {
                'success': False,
                'message': 'Missing required columns in holiday import file.',
                'errors': [f"Missing columns: {', '.join(missing_headers)}"],
            }

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[header_map['date']] or not row[header_map['holiday']]:
                continue

            try:
                date_value = row[header_map['date']]
                if isinstance(date_value, datetime):
                    holiday_date = date_value.date()
                elif hasattr(date_value, 'year') and hasattr(date_value, 'month') and hasattr(date_value, 'day'):
                    holiday_date = date_value
                else:
                    holiday_date = datetime.strptime(str(date_value).strip(), '%Y-%m-%d').date()

                holiday_name = str(row[header_map['holiday']]).strip()
                location = str(row[header_map['applicability']]).strip() if row[header_map['applicability']] else ''

                holiday, created = Holiday.objects.update_or_create(
                    date=holiday_date,
                    name=holiday_name,
                    location=location,
                    defaults={
                        'holiday_type': 'PUBLIC_HOLIDAY',
                    }
                )

                if created:
                    results['created_count'] += 1
                else:
                    results['updated_count'] += 1
            except Exception as exc:
                results['errors'].append(f'Row {row_idx}: {str(exc)}')

        workbook.close()

        if results['errors']:
            results['success'] = False
            results['message'] = f"Holiday import completed with {len(results['errors'])} errors"

        return results
    except Exception as exc:
        import_logger.error(f"Failed to import holiday file: {str(exc)}", exc_info=True)
        return {
            'success': False,
            'message': f'Failed to import holiday file: {str(exc)}',
            'errors': [str(exc)],
        }


def import_timesheet_file(file_path, import_date=None):
    """
    Import timesheet data from XLSX file using bulk_create for performance optimization.
    Expected sheets: Timesheet Summary, Timesheet Details, Attendance Summary, Attendance Details
    
    Uses bulk_create() for batch insert operations to reduce database I/O calls significantly.
    Batch size of 1000 records per insert to balance memory usage and performance.
    
    Args:
        file_path: Path to the XLSX file
        import_date: Date/Month of the import (optional, for filtering later)
    """
    import_logger.info(f"Starting timesheet import from file: {file_path}, import_date: {import_date}")
    
    try:
        workbook = load_workbook(file_path)
        results = {
            'success': True,
            'message': 'Import completed successfully',
            'summary_count': 0,
            'details_count': 0,
            'attendance_summary_count': 0,
            'attendance_details_count': 0,
            'errors': []
        }
        
        import_logger.info(f"Excel file loaded. Sheets found: {workbook.sheetnames}")
        BATCH_SIZE = 1000  # Process records in batches of 1000

        # Import Timesheet Summary
        if 'Timesheet Summary' in workbook.sheetnames:
            try:
                ws = workbook['Timesheet Summary']
                summaries_to_create = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Skip header
                        continue
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    try:
                        summary = TimesheetSummary(
                            email_id=str(row[0]).strip(),
                            team_member=str(row[1]).strip(),
                            project_id=int(row[2]),
                            project=str(row[3]).strip(),
                            total_hours=Decimal(str(row[4])),
                            import_date=import_date
                        )
                        summaries_to_create.append(summary)
                    except Exception as e:
                        results['errors'].append(f"Timesheet Summary Row {row_idx}: {str(e)}")
                
                # Bulk insert in batches
                if summaries_to_create:
                    TimesheetSummary.objects.bulk_create(summaries_to_create, batch_size=BATCH_SIZE)
                    results['summary_count'] = len(summaries_to_create)
                    import_logger.info(f"Bulk created {results['summary_count']} Timesheet Summary records")
                    
            except Exception as e:
                results['errors'].append(f"Error processing Timesheet Summary: {str(e)}")

        # Import Timesheet Details
        if 'Timesheet Details' in workbook.sheetnames:
            try:
                ws = workbook['Timesheet Details']
                details_to_create = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Skip header
                        continue
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    try:
                        # Parse date
                        date_val = row[4]
                        if isinstance(date_val, datetime):
                            date_val = date_val.date()
                        else:
                            date_val = datetime.strptime(str(date_val), '%Y-%m-%d').date()
                        
                        detail = TimesheetDetails(
                            email_id=str(row[0]).strip(),
                            team_member=str(row[1]).strip(),
                            project_id=int(row[2]),
                            project=str(row[3]).strip(),
                            date=date_val,
                            time_logged=Decimal(str(row[5])),
                            comment=str(row[6]).strip() if row[6] else '',
                            import_date=import_date
                        )
                        details_to_create.append(detail)
                    except Exception as e:
                        results['errors'].append(f"Timesheet Details Row {row_idx}: {str(e)}")
                
                # Bulk insert in batches
                if details_to_create:
                    TimesheetDetails.objects.bulk_create(details_to_create, batch_size=BATCH_SIZE)
                    results['details_count'] = len(details_to_create)
                    import_logger.info(f"Bulk created {results['details_count']} Timesheet Details records")
                    
            except Exception as e:
                results['errors'].append(f"Error processing Timesheet Details: {str(e)}")

        # Import Attendance Summary
        if 'Attendance Summary' in workbook.sheetnames:
            try:
                ws = workbook['Attendance Summary']
                att_summaries_to_create = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Skip header
                        continue
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    try:
                        summary = AttendanceSummary(
                            email_id=str(row[0]).strip(),
                            team_member=str(row[1]).strip(),
                            business_hours=Decimal(str(row[2])),
                            available_hours=Decimal(str(row[3])),
                            project_hours=Decimal(str(row[4])),
                            import_date=import_date
                        )
                        att_summaries_to_create.append(summary)
                    except Exception as e:
                        results['errors'].append(f"Attendance Summary Row {row_idx}: {str(e)}")
                
                # Bulk insert in batches
                if att_summaries_to_create:
                    AttendanceSummary.objects.bulk_create(att_summaries_to_create, batch_size=BATCH_SIZE)
                    results['attendance_summary_count'] = len(att_summaries_to_create)
                    import_logger.info(f"Bulk created {results['attendance_summary_count']} Attendance Summary records")
                    
            except Exception as e:
                results['errors'].append(f"Error processing Attendance Summary: {str(e)}")

        # Import Attendance Details
        if 'Attendance Details' in workbook.sheetnames:
            try:
                ws = workbook['Attendance Details']
                att_details_to_create = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Skip header
                        continue
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    try:
                        # Parse date
                        date_val = row[2]
                        if isinstance(date_val, datetime):
                            date_val = date_val.date()
                        else:
                            date_val = datetime.strptime(str(date_val), '%Y-%m-%d').date()
                        
                        detail = AttendanceDetails(
                            email_id=str(row[0]).strip(),
                            team_member=str(row[1]).strip(),
                            date=date_val,
                            business_hours=Decimal(str(row[3])),
                            available_hours=Decimal(str(row[4])),
                            remarks=str(row[5]).strip() if row[5] else '',
                            import_date=import_date
                        )
                        att_details_to_create.append(detail)
                    except Exception as e:
                        results['errors'].append(f"Attendance Details Row {row_idx}: {str(e)}")
                
                # Bulk insert in batches
                if att_details_to_create:
                    AttendanceDetails.objects.bulk_create(att_details_to_create, batch_size=BATCH_SIZE)
                    results['attendance_details_count'] = len(att_details_to_create)
                    import_logger.info(f"Bulk created {results['attendance_details_count']} Attendance Details records")
                    
            except Exception as e:
                results['errors'].append(f"Error processing Attendance Details: {str(e)}")

        if results['errors']:
            results['success'] = False
            results['message'] = f"Import completed with {len(results['errors'])} errors"
            import_logger.warning(f"Import completed with {len(results['errors'])} errors: {results['errors']}")
        else:
            results['message'] = 'Import completed successfully'
            import_logger.info(f"Import completed successfully (using bulk_create). Summary={results['summary_count']}, Details={results['details_count']}, "
                              f"AttSummary={results['attendance_summary_count']}, AttDetails={results['attendance_details_count']}")
        
        workbook.close()
        return results

    except Exception as e:
        import_logger.error(f"Failed to import file: {str(e)}", exc_info=True)
        return {
            'success': False,
            'message': f"Failed to import file: {str(e)}",
            'errors': [str(e)]
        }


def import_client_timesheet_file(file_path, import_date=None, project_name=''):
    """
    Import client timesheet data from XLSX file and create/update TimesheetEntry records in DRAFT status.
    
    Expected format:
    - Row 2: Date | Employee Name 1 | Employee Name 2 | ... | Employee Name N
    - Row 3+: Date | Hours for Employee 1 | Hours for Employee 2 | ... | Hours for Employee N
    - Data spans from start to end of month (including weekends/holidays)
    - Only create entries for cells with value > 0
    
    Args:
        file_path: Path to the XLSX file
        import_date: Date/Month of the import (month start date)
        project_name: Project name to assign to all imported entries
    
    Returns:
        Dictionary with success status, message, and counts
    """
    import_logger.info(f"Starting client timesheet import from file: {file_path}, import_date: {import_date}, project: {project_name}")
    
    try:
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        
        results = {
            'success': True,
            'message': 'Import completed successfully',
            'entries_created': 0,
            'entries_updated': 0,
            'entries_skipped': 0,
            'errors': [],
            'skipped_details': [],
            'failed_details': [],
        }
        
        if not project_name or project_name.strip() == '':
            results['errors'].append('Project name is required')
            _record_client_import_issue(
                results,
                'failed_details',
                reason='Project name is required',
                project=project_name,
            )
            results['success'] = False
            import_logger.error("Client timesheet import failed: Project name not provided")
            return results
        
        project_name = project_name.strip()
        
        # Row 2 contains headers: Date | Employee Name 1 | Employee Name 2 | ...
        header_row = None
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), 2):
            header_row = row
            break
        
        if not header_row:
            results['errors'].append('Header row (Row 2) not found')
            _record_client_import_issue(
                results,
                'failed_details',
                row=2,
                reason='Header row (Row 2) not found',
            )
            results['success'] = False
            return results
        
        # Extract employee names from header (skip first column which is Date).
        # Some client exports put an external/client id on a second line in the
        # header cell, so keep the raw value for messages and match on line one.
        employee_names = []
        employee_header_values = []
        employee_indices = []
        for idx, col_value in enumerate(header_row):
            if idx > 0 and col_value:  # Skip Date column (index 0)
                employee_header_values.append(str(col_value).strip())
                employee_names.append(_employee_name_from_header(col_value))
                employee_indices.append(idx)
        
        if not employee_names:
            results['errors'].append('No employee names found in header row')
            _record_client_import_issue(
                results,
                'failed_details',
                row=2,
                reason='No employee names found in header row',
            )
            results['success'] = False
            return results
        
        import_logger.info(f"Found {len(employee_names)} employees in header: {', '.join(employee_names)}")
        
        # Build employee lookup: normalized name -> Employee object
        employees_by_name = {
            _normalize_name(employee.name).casefold(): employee
            for employee in Employee.objects.select_related('project').all()
        }
        employee_lookup = {}
        for idx, emp_name in enumerate(employee_names):
            employee = employees_by_name.get(_normalize_name(emp_name).casefold())
            employee_lookup[emp_name] = employee
            if not employee:
                raw_header = employee_header_values[idx]
                column = _column_letter(employee_indices[idx])
                if raw_header != emp_name:
                    error_message = f"Employee '{emp_name}' not found in database (header: '{raw_header}')"
                else:
                    error_message = f"Employee '{emp_name}' not found in database"
                results['errors'].append(error_message)
                _record_client_import_issue(
                    results,
                    'failed_details',
                    row=2,
                    column=column,
                    employee=emp_name,
                    raw_header=raw_header,
                    reason=error_message,
                    project=project_name,
                )
        
        # Row 3+ contains data: Date | Hours for Employee 1 | Hours for Employee 2 | ...
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
            if not row or not row[0]:  # Skip empty rows
                continue
            
            try:
                # Parse date from first column
                date_val = row[0]
                if isinstance(date_val, datetime):
                    entry_date = date_val.date()
                elif isinstance(date_val, date):
                    entry_date = date_val
                else:
                    entry_date = datetime.strptime(str(date_val).strip(), '%Y-%m-%d').date()
                
                # Process each employee's hours for this date
                for idx, emp_idx in enumerate(employee_indices):
                    emp_name = employee_names[idx]
                    raw_header = employee_header_values[idx]
                    column = _column_letter(emp_idx)

                    # Get hours value
                    try:
                        hours_value = row[emp_idx] if emp_idx < len(row) else None
                        if hours_value is None or hours_value == '':
                            # Empty cell, skip
                            results['entries_skipped'] += 1
                            _record_client_import_issue(
                                results,
                                'skipped_details',
                                row=row_idx,
                                column=column,
                                date=entry_date,
                                employee=emp_name,
                                raw_header=raw_header,
                                value=hours_value,
                                reason='Empty hours cell',
                                project=project_name,
                            )
                            continue
                        
                        hours = float(hours_value)
                        
                        # Only create if hours > 0
                        if hours <= 0:
                            results['entries_skipped'] += 1
                            _record_client_import_issue(
                                results,
                                'skipped_details',
                                row=row_idx,
                                column=column,
                                date=entry_date,
                                employee=emp_name,
                                raw_header=raw_header,
                                value=hours_value,
                                reason='Hours value is zero or negative',
                                project=project_name,
                            )
                            continue

                        employee = employee_lookup.get(emp_name)
                        if not employee:
                            results['entries_skipped'] += 1
                            _record_client_import_issue(
                                results,
                                'skipped_details',
                                row=row_idx,
                                column=column,
                                date=entry_date,
                                employee=emp_name,
                                raw_header=raw_header,
                                value=hours_value,
                                reason='Employee not found in database',
                                project=project_name,
                            )
                            continue
                        
                        entry_project = employee.project.project if employee.project else project_name

                        # Prefer updating an existing draft for the employee/date.
                        # Previous imports could have used a different project text,
                        # but the Client Timesheet page treats employee/date as the
                        # user-facing row to correct for a day.
                        existing_entries = TimesheetEntry.objects.filter(
                            employee=employee,
                            date=entry_date,
                        ).order_by('project', 'id')

                        entry = existing_entries.filter(project=entry_project).first()
                        if not entry:
                            entry = existing_entries.filter(status='DRAFT').first()

                        if entry:
                            # Update existing entry if in DRAFT status
                            if entry.status == 'DRAFT':
                                entry.project = entry_project
                                entry.hours = Decimal(str(hours))
                                entry.comments = f'Updated from client timesheet import on {datetime.now().strftime("%Y-%m-%d %H:%M")}'
                                entry.save()
                                results['entries_updated'] += 1
                                import_logger.info(f"Updated entry: {employee.name} on {entry_date} - {hours} hours")
                            else:
                                # Cannot update submitted entries
                                results['entries_skipped'] += 1
                                _record_client_import_issue(
                                    results,
                                    'skipped_details',
                                    row=row_idx,
                                    column=column,
                                    date=entry_date,
                                    employee=emp_name,
                                    raw_header=raw_header,
                                    value=hours_value,
                                    reason='Existing entry is submitted and cannot be updated',
                                    project=entry_project,
                                )
                                import_logger.warning(f"Skipped submitted entry: {employee.name} on {entry_date}")
                        else:
                            TimesheetEntry.objects.create(
                                employee=employee,
                                date=entry_date,
                                project=entry_project,
                                hours=Decimal(str(hours)),
                                status='DRAFT',
                                comments=f'Imported from client timesheet on {datetime.now().strftime("%Y-%m-%d %H:%M")}'
                            )
                            results['entries_created'] += 1
                            import_logger.info(f"Created entry: {employee.name} on {entry_date} - {hours} hours")
                    
                    except (ValueError, TypeError) as e:
                        # Invalid hours value
                        results['entries_skipped'] += 1
                        _record_client_import_issue(
                            results,
                            'failed_details',
                            row=row_idx,
                            column=column,
                            date=entry_date,
                            employee=emp_name,
                            raw_header=raw_header,
                            value=hours_value if 'hours_value' in locals() else '',
                            reason=f'Invalid hours value: {str(e)}',
                            project=project_name,
                        )
                        import_logger.warning(f"Invalid hours value for {emp_name} on {entry_date}: {row[emp_idx]}")
                        continue
            
            except Exception as e:
                results['errors'].append(f"Row {row_idx}: {str(e)}")
                _record_client_import_issue(
                    results,
                    'failed_details',
                    row=row_idx,
                    value=row[0] if row else '',
                    reason=f'Row failed: {str(e)}',
                    project=project_name,
                )
                import_logger.warning(f"Error processing row {row_idx}: {str(e)}")
                continue
        
        if results['errors']:
            results['message'] = f"Import completed with {len(results['errors'])} warnings"
            import_logger.warning(f"Client timesheet import completed with warnings: {results['errors']}")
        else:
            results['message'] = 'Import completed successfully'
        
        import_logger.info(f"Client timesheet import completed. Created={results['entries_created']}, "
                          f"Updated={results['entries_updated']}, Skipped={results['entries_skipped']}")
        if results['skipped_details']:
            import_logger.warning(f"Client timesheet skipped details: {results['skipped_details']}")
        if results['failed_details']:
            import_logger.error(f"Client timesheet failed details: {results['failed_details']}")
        
        workbook.close()
        return results
    
    except Exception as e:
        import_logger.error(f"Failed to import client timesheet file: {str(e)}", exc_info=True)
        return {
            'success': False,
            'message': f"Failed to import file: {str(e)}",
            'entries_created': 0,
            'entries_updated': 0,
            'entries_skipped': 0,
            'errors': [str(e)],
            'skipped_details': [],
            'failed_details': [{
                'row': '',
                'column': '',
                'date': '',
                'employee': '',
                'raw_header': '',
                'value': '',
                'reason': str(e),
                'project': project_name,
            }]
        }


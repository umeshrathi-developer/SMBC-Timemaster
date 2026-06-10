"""Utility functions for timesheet and holiday import"""
import csv
import os
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
import logging
from django.conf import settings
from django.contrib.auth.models import Group, User
from django.db import transaction
from django.db.models import Q
from .models import (
    TimesheetSummary, TimesheetDetails, AttendanceSummary, AttendanceDetails,
    Holiday, Employee, Project, Location, TimesheetEntry, CompOff, Accrual
)

import_logger = logging.getLogger('timesheet.import')


def _normalize_header(value):
    """Normalize Excel headers so small naming variations are accepted."""
    return str(value or '').strip().lower().replace(' ', '_').replace('-', '_')


def _normalize_lookup(value):
    """Normalize employee identifiers for case-insensitive lookup."""
    return str(value or '').strip().lower()


def _parse_excel_date(value):
    """Parse a date value from Excel or a supported string format."""
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except (TypeError, ValueError):
            pass

    value = str(value or '').strip()
    for date_format in (
        '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y',
        '%d-%b-%Y', '%d-%B-%Y'
    ):
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    raise ValueError('Invalid date. Use YYYY-MM-DD or DD-MM-YYYY.')


def _parse_hours(value):
    """Parse an Excel hour cell and return Decimal hours."""
    if value in (None, ''):
        return Decimal('0')
    try:
        return Decimal(str(value).strip().replace(',', ''))
    except (InvalidOperation, AttributeError):
        raise ValueError(f'Invalid hours value "{value}".')


def _format_import_value(value):
    """Format imported values for display in exception reports."""
    if value is None:
        return ''
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return str(value)


def _record_client_import_issue(
    results, issue_type, row, column, employee_header='', employee=None,
    entry_date=None, hours='', reason=''
):
    """Record a skipped or failed import cell with enough detail to fix the file."""
    results[f'{issue_type}_records'].append({
        'type': issue_type.upper(),
        'row': row,
        'column': column,
        'employee': employee.name if employee else '',
        'employee_header': employee_header,
        'date': _format_import_value(entry_date),
        'hours': _format_import_value(hours),
        'reason': reason,
    })


def _write_client_import_exception_report(results):
    """Write skipped and failed import details to a CSV in the workspace."""
    records = results.get('failed_records', []) + results.get('skipped_records', [])
    if not records:
        return ''

    reports_dir = os.path.join(settings.BASE_DIR, 'import_reports')
    os.makedirs(reports_dir, exist_ok=True)
    filename = f"client_timesheet_import_exceptions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    report_path = os.path.join(reports_dir, filename)

    with open(report_path, 'w', newline='', encoding='utf-8') as report_file:
        writer = csv.DictWriter(
            report_file,
            fieldnames=['type', 'row', 'column', 'employee', 'employee_header', 'date', 'hours', 'reason']
        )
        writer.writeheader()
        writer.writerows(records)

    return os.path.relpath(report_path, settings.BASE_DIR)


def _get_client_timesheet_import_status():
    """Return the configured status for imported client timesheet entries."""
    status = getattr(settings, 'CLIENT_TIMESHEET_IMPORT_STATUS', 'DRAFT')
    status = str(status).strip().upper()
    valid_statuses = {choice[0] for choice in TimesheetEntry.STATUS_CHOICES}
    if status not in valid_statuses:
        raise ValueError(
            'CLIENT_TIMESHEET_IMPORT_STATUS must be one of: '
            f"{', '.join(sorted(valid_statuses))}."
        )
    return status


def _get_employee_location_name(employee):
    """Return an employee location name for holiday matching."""
    location = getattr(employee, 'location', None)
    return getattr(location, 'name', location) or ''


def _holiday_applies_to_location(holiday, location):
    """Return whether a holiday applies to a location.

    Holiday.location can be blank for global holidays or comma-separated for
    multiple applicable locations.
    """
    holiday_location = str(getattr(holiday, 'location', '') or '').strip()
    if not holiday_location or not location:
        return True

    location = str(location).strip().lower()
    holiday_locations = [
        item.strip().lower()
        for item in holiday_location.split(',')
        if item.strip()
    ]
    return location in holiday_locations


def _is_weekend_or_public_holiday(entry_date, employee):
    """Return whether an imported entry is eligible for CompOff accrual."""
    if entry_date.weekday() >= 5:
        return True

    holidays = Holiday.objects.filter(
        date=entry_date,
        holiday_type__in=['PUBLIC_HOLIDAY']
    )
    location = _get_employee_location_name(employee)

    return any(_holiday_applies_to_location(holiday, location) for holiday in holidays)


def _is_weekend_or_holiday_for_accrual(entry_date, employee):
    """Return whether an imported entry should create an Accrual record."""
    if entry_date.weekday() >= 5:
        return True

    holidays = Holiday.objects.filter(
        date=entry_date,
        holiday_type__in=['PUBLIC_HOLIDAY', 'US_HOLIDAY']
    )
    location = _get_employee_location_name(employee)

    return any(_holiday_applies_to_location(holiday, location) for holiday in holidays)


def _create_import_compoff_if_eligible(entry, results):
    """Create a pending CompOff for eligible imported submitted entries."""
    if entry.status != 'SUBMITTED':
        return
    if entry.hours < 8:
        return
    if not _is_weekend_or_public_holiday(entry.date, entry.employee):
        return
    if CompOff.objects.filter(employee=entry.employee, working_date=entry.date).exists():
        return

    CompOff.objects.create(
        employee=entry.employee,
        working_date=entry.date,
        compoff_date=None,
        status='PENDING',
        notes=f'Auto-generated Comp-Off from client timesheet import ({entry.date.strftime("%a, %b %d, %Y")})'
    )
    results['compoff_created_count'] += 1


def _create_import_accrual_if_eligible(entry, results):
    """Create a pending Accrual for eligible imported submitted entries."""
    if entry.status != 'SUBMITTED':
        return
    if entry.hours < 8:
        return
    if not _is_weekend_or_holiday_for_accrual(entry.date, entry.employee):
        return
    if Accrual.objects.filter(
        employee=entry.employee,
        working_date=entry.date,
        adjusted_date__isnull=True,
    ).exists():
        return

    Accrual.objects.create(
        employee=entry.employee,
        working_date=entry.date,
        adjusted_date=None,
        adjustment_reason=None,
        status='PENDING',
        notes=f'Auto-generated Accrual from client timesheet import ({entry.date.strftime("%a, %b %d, %Y")})'
    )
    results['accrual_created_count'] = results.get('accrual_created_count', 0) + 1


def _delete_pending_import_compoff_if_ineligible(entry, results):
    """Remove an unused pending CompOff when an authoritative import clears eligibility."""
    if entry.status == 'SUBMITTED' and entry.hours >= 8 and _is_weekend_or_public_holiday(entry.date, entry.employee):
        return

    deleted_count = CompOff.objects.filter(
        employee=entry.employee,
        working_date=entry.date,
        status='PENDING',
        compoff_date__isnull=True,
    ).delete()[0]
    results['compoff_deleted_count'] += deleted_count


def _sync_import_compoff(entry, results):
    """Keep pending CompOffs and Accruals aligned with the imported cell value."""
    _delete_pending_import_compoff_if_ineligible(entry, results)
    _create_import_compoff_if_eligible(entry, results)
    _create_import_accrual_if_eligible(entry, results)


def _is_import_workday_for_compoff_use(entry_date, employee):
    """Return whether a date can use a pending CompOff in lieu of hours."""
    if entry_date.weekday() >= 5:
        return False

    location = _get_employee_location_name(employee)
    holidays = Holiday.objects.filter(
        date=entry_date,
        holiday_type__in=['PUBLIC_HOLIDAY', 'US_HOLIDAY', 'SPECIAL_HOLIDAY']
    )
    return not any(_holiday_applies_to_location(holiday, location) for holiday in holidays)


def _has_positive_submitted_hours(employee, entry_date):
    """Return whether an employee has submitted positive hours for a date."""
    return TimesheetEntry.objects.filter(
        employee=employee,
        date=entry_date,
        status='SUBMITTED',
        hours__gt=0,
    ).exists()


def _get_import_candidate_dates(results):
    """Return employee/date pairs to evaluate during import adjustment logic."""
    affected = set(results.get('_affected_employee_dates', set()))
    if not affected:
        return set()

    start_date = results.get('import_start_date')
    end_date = results.get('import_end_date')
    if start_date and end_date:
        employee_ids = {employee_id for employee_id, _ in affected}
        current = start_date
        candidate_dates = set()
        while current <= end_date:
            candidate_dates.update((employee_id, current) for employee_id in employee_ids)
            current += timedelta(days=1)
        return candidate_dates

    return affected


def _auto_adjust_pending_accruals_for_import(results):
    """Mark pending Accruals as adjusted for imported leave/public-holiday-off days."""
    if results.get('import_status') != 'SUBMITTED':
        return

    affected = _get_import_candidate_dates(results)
    if not affected:
        return

    employee_ids = {employee_id for employee_id, _ in affected}
    employees = {
        employee.id: employee
        for employee in Employee.objects.filter(id__in=employee_ids).select_related('location')
    }

    for employee_id, entry_date in sorted(affected, key=lambda item: (item[0], item[1])):
        employee = employees.get(employee_id)
        if not employee:
            continue

        if not _has_positive_submitted_hours(employee, entry_date):
            # Leave taken on a regular weekday: reduce pending accruals
            if entry_date.weekday() < 5 and not any(
                holiday.date == entry_date and _holiday_applies_to_location(holiday, _get_employee_location_name(employee))
                for holiday in Holiday.objects.filter(date=entry_date, holiday_type__in=['PUBLIC_HOLIDAY', 'US_HOLIDAY', 'SPECIAL_HOLIDAY'])
            ):
                existing = Accrual.objects.filter(
                    employee=employee,
                    status='ADJUSTED',
                    adjusted_date=entry_date,
                    adjustment_reason='LEAVE_TAKEN',
                ).exists()
                if not existing:
                    oldest_accrual = Accrual.objects.filter(
                        employee=employee,
                        status='PENDING',
                        adjusted_date__isnull=True,
                        working_date__lte=entry_date,
                    ).order_by('created_date').first()
                    if oldest_accrual:
                        oldest_accrual.adjusted_date = entry_date
                        oldest_accrual.adjustment_reason = 'LEAVE_TAKEN'
                        oldest_accrual.save()
                        results['accrual_adjusted_count'] = results.get('accrual_adjusted_count', 0) + 1

            # Public holiday off: reduce pending accruals on the holiday date itself
            public_holidays = Holiday.objects.filter(
                date=entry_date,
                holiday_type='PUBLIC_HOLIDAY',
            )
            if any(_holiday_applies_to_location(holiday, _get_employee_location_name(employee)) for holiday in public_holidays):
                existing = Accrual.objects.filter(
                    employee=employee,
                    status='ADJUSTED',
                    adjusted_date=entry_date,
                    adjustment_reason='PUBLIC_HOLIDAY_OFF',
                ).exists()
                if not existing:
                    oldest_accrual = Accrual.objects.filter(
                        employee=employee,
                        status='PENDING',
                        adjusted_date__isnull=True,
                        working_date__lte=entry_date,
                    ).order_by('created_date').first()
                    if oldest_accrual:
                        oldest_accrual.adjusted_date = entry_date
                        oldest_accrual.adjustment_reason = 'PUBLIC_HOLIDAY_OFF'
                        oldest_accrual.save()
                        results['accrual_adjusted_count'] = results.get('accrual_adjusted_count', 0) + 1


def _auto_apply_pending_compoffs_for_import(results):
    """Mark pending CompOffs as taken for imported weekday zero/blank days."""
    if results.get('import_status') != 'SUBMITTED':
        return

    affected = _get_import_candidate_dates(results)
    if not affected:
        return

    employee_ids = {employee_id for employee_id, _ in affected}
    employees = {
        employee.id: employee
        for employee in Employee.objects.filter(id__in=employee_ids).select_related('location')
    }

    for employee_id, entry_date in sorted(affected, key=lambda item: (item[0], item[1])):
        employee = employees.get(employee_id)
        if not employee:
            continue
        if not _is_import_workday_for_compoff_use(entry_date, employee):
            continue
        if _has_positive_submitted_hours(employee, entry_date):
            continue
        if CompOff.objects.filter(employee=employee, status='TAKEN', compoff_date=entry_date).exists():
            continue

        oldest_compoff = CompOff.objects.filter(
            employee=employee,
            status='PENDING',
            compoff_date__isnull=True,
            working_date__lte=entry_date,
        ).order_by('created_date').first()
        if oldest_compoff:
            oldest_compoff.compoff_date = entry_date
            oldest_compoff.status = 'TAKEN'
            oldest_compoff.save()
            results['compoff_taken_count'] += 1


def _build_employee_lookup():
    """Build a lookup for employee id, email, and display name headers."""
    lookup = {}
    for employee in Employee.objects.filter(is_active=True).select_related('project'):
        keys = [
            employee.employee_id,
            employee.email,
            employee.name,
            f'{employee.name} ({employee.employee_id})',
        ]
        for key in keys:
            normalized_key = _normalize_lookup(key)
            if normalized_key and normalized_key not in lookup:
                lookup[normalized_key] = employee
    return lookup


def _find_employee(value, employee_lookup):
    """Find an active employee from a spreadsheet header or identifier cell."""
    lookup_value = str(value or '').strip()
    if not lookup_value:
        return None

    candidates = [
        lookup_value,
        lookup_value.splitlines()[0].strip(),
    ]
    if '(' in lookup_value and ')' in lookup_value:
        candidates.append(lookup_value[lookup_value.rfind('(') + 1:lookup_value.rfind(')')].strip())

    for candidate in candidates:
        employee = employee_lookup.get(_normalize_lookup(candidate))
        if employee:
            return employee
    return None


def _split_employee_name(name):
    """Split an employee name into first and last names for User records."""
    parts = str(name or '').strip().split()
    if not parts:
        return '', ''
    if len(parts) == 1:
        return parts[0], ''
    return parts[0], parts[-1]


def _build_employee_username(name):
    """Build username in firstname.lastname format."""
    first_name, last_name = _split_employee_name(name)
    username = first_name.lower()
    if last_name:
        username = f'{username}.{last_name.lower()}'
    return username.replace(' ', '.')


def _parse_bool(value):
    """Parse common Excel boolean values."""
    if isinstance(value, bool):
        return value
    if value in (None, ''):
        return True

    normalized_value = str(value).strip().lower()
    if normalized_value in {'true', 'yes', 'y', '1', 'active'}:
        return True
    if normalized_value in {'false', 'no', 'n', '0', 'inactive'}:
        return False
    raise ValueError(f'Invalid is_active value "{value}". Use TRUE/FALSE or YES/NO.')


def _get_next_project_id():
    """Return the next integer project id."""
    return (Project.objects.order_by('-project_id').values_list('project_id', flat=True).first() or 0) + 1


def _get_or_create_import_project(project_name):
    """Find or create a project from employee import data."""
    project_name = str(project_name or '').strip()
    if not project_name:
        return None

    project = Project.objects.filter(project__iexact=project_name).first()
    if project:
        return project

    return Project.objects.create(
        project_id=_get_next_project_id(),
        department_name='Risk Tech',
        project=project_name,
        project_code=1000,
        manager=project_name,
    )


def _get_or_create_location(location_name):
    """Find or create a location from employee import data."""
    location_name = str(location_name or '').strip()
    if not location_name:
        return None

    location = Location.objects.filter(name__iexact=location_name).first()
    if location:
        return location

    return Location.objects.create(name=location_name)


def _create_or_update_employee_user(employee):
    """Create or update the linked Django user for an employee."""
    if not employee.name:
        return None, False

    first_name, last_name = _split_employee_name(employee.name)
    base_username = _build_employee_username(employee.name)
    if not base_username:
        base_username = employee.employee_id.lower()

    user_created = False
    user = employee.user
    if not user:
        user = User.objects.filter(username=base_username).first()
        if user and hasattr(user, 'employee') and user.employee != employee:
            user = None

    if not user:
        username = base_username
        suffix = 1
        while User.objects.filter(username=username).exists():
            username = f'{base_username}{suffix}'
            suffix += 1
        user = User.objects.create_user(
            username=username,
            email=employee.email,
            password=username,
            first_name=first_name,
            last_name=last_name,
        )
        user_created = True

    user.email = employee.email
    user.first_name = first_name
    user.last_name = last_name
    user.is_active = employee.is_active
    user.save()

    group, _ = Group.objects.get_or_create(name='Employee')
    user.groups.add(group)

    if employee.user_id != user.id:
        employee.user = user
        employee.save(update_fields=['user', 'updated_date'])

    return user, user_created


def import_employee_file(file_obj):
    """
    Import employee master data from an Excel file.

    Expected columns:
    - name
    - employee_id
    - email
    - role (optional: Dev or QA; defaults to Dev)
    - project
    - location
    - is_active
    """
    workbook = None
    try:
        workbook = load_workbook(file_obj, data_only=True)
        worksheet = workbook.active
        results = {
            'success': True,
            'message': 'Employee import completed successfully',
            'created_count': 0,
            'updated_count': 0,
            'user_created_count': 0,
            'project_created_count': 0,
            'errors': [],
        }

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {
                'success': False,
                'message': 'The uploaded file is empty.',
                'created_count': 0,
                'updated_count': 0,
                'user_created_count': 0,
                'project_created_count': 0,
                'errors': ['Missing header row.'],
            }

        header_map = {
            _normalize_header(column): index
            for index, column in enumerate(header_row)
            if column is not None
        }
        required_headers = ['name', 'employee_id', 'email', 'project', 'location', 'is_active']
        missing_headers = [header for header in required_headers if header not in header_map]
        if missing_headers:
            return {
                'success': False,
                'message': 'Missing required columns in employee import file.',
                'created_count': 0,
                'updated_count': 0,
                'user_created_count': 0,
                'project_created_count': 0,
                'errors': [f"Missing columns: {', '.join(missing_headers)}"],
            }

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue

            try:
                name = str(row[header_map['name']] or '').strip()
                employee_id = str(row[header_map['employee_id']] or '').strip()
                email = str(row[header_map['email']] or '').strip()
                role = str(row[header_map['role']] or 'Dev').strip() if 'role' in header_map else 'Dev'
                project_name = str(row[header_map['project']] or '').strip()
                location_name = str(row[header_map['location']] or '').strip()
                is_active = _parse_bool(row[header_map['is_active']])

                if not name:
                    raise ValueError('name is required.')
                if not employee_id:
                    raise ValueError('employee_id is required.')
                if not email:
                    raise ValueError('email is required.')
                if not project_name:
                    raise ValueError('project is required.')
                if role not in {'Dev', 'QA'}:
                    raise ValueError('role must be Dev or QA.')

                with transaction.atomic():
                    existing_project_count = Project.objects.count()
                    project = _get_or_create_import_project(project_name)
                    if Project.objects.count() > existing_project_count:
                        results['project_created_count'] += 1
                    location = _get_or_create_location(location_name)

                    employee, created = Employee.objects.update_or_create(
                        employee_id=employee_id,
                        defaults={
                            'name': name,
                            'email': email,
                            'role': role,
                            'project': project,
                            'location': location,
                            'is_active': is_active,
                        }
                    )
                    _, user_created = _create_or_update_employee_user(employee)

                if created:
                    results['created_count'] += 1
                else:
                    results['updated_count'] += 1
                if user_created:
                    results['user_created_count'] += 1
            except Exception as exc:
                results['errors'].append(f'Row {row_idx}: {str(exc)}')

        if results['errors']:
            results['success'] = False
            results['message'] = f"Employee import completed with {len(results['errors'])} errors"

        return results
    except Exception as exc:
        import_logger.error(f"Failed to import employee file: {str(exc)}", exc_info=True)
        return {
            'success': False,
            'message': f'Failed to import employee file: {str(exc)}',
            'created_count': 0,
            'updated_count': 0,
            'user_created_count': 0,
            'project_created_count': 0,
            'errors': [str(exc)],
        }
    finally:
        if workbook:
            workbook.close()


def _save_client_timesheet_entry(
    employee, entry_date, hours, overwrite_drafts, results, comments='',
    row_idx='', column_letter='', employee_header='', raw_hours=''
):
    """Create or update one imported client timesheet cell.

    Client imports are authoritative for every employee/date cell they contain,
    including blank and zero-hour cells.
    """
    if hours < 0:
        raise ValueError('Hours must be between 0 and 12.')
    if hours > 12:
        raise ValueError('Hours must be between 0 and 12.')
    if not employee.project:
        raise ValueError(f'Employee "{employee.name}" has no assigned project.')

    project = employee.project.project
    existing_entry = TimesheetEntry.objects.filter(
        employee=employee,
        date=entry_date,
        project=project
    ).first()

    if existing_entry:
        existing_entry.hours = hours
        existing_entry.comments = comments
        existing_entry.status = results['import_status']
        existing_entry.save()
        _sync_import_compoff(existing_entry, results)
        results['_affected_employee_dates'].add((existing_entry.employee_id, existing_entry.date))
        results['updated_count'] += 1
        return

    entry = TimesheetEntry.objects.create(
        employee=employee,
        date=entry_date,
        project=project,
        hours=hours,
        comments=comments,
        status=results['import_status']
    )
    _sync_import_compoff(entry, results)
    results['_affected_employee_dates'].add((entry.employee_id, entry.date))
    results['created_count'] += 1


def _import_client_timesheet_matrix(
    worksheet, header_row, results, overwrite_drafts, data_start_row=2, header_row_number=1,
    start_date=None, end_date=None
):
    """Import Date x Employee hour matrix: Date | Employee A | Employee B | ..."""
    employee_lookup = _build_employee_lookup()
    date_column_index = None
    employee_columns = []

    for index, header in enumerate(header_row):
        normalized_header = _normalize_header(header)
        if normalized_header in {'date', 'work_date', 'day'}:
            date_column_index = index
            continue

        employee = _find_employee(header, employee_lookup)
        if employee:
            employee_columns.append((index, employee, str(header).strip()))
        elif header not in (None, ''):
            results['errors'].append(
                f'Row {header_row_number}, column {get_column_letter(index + 1)}: '
                f'Active employee not found for header "{str(header).strip()}".'
            )
            _record_client_import_issue(
                results, 'failed', header_row_number, get_column_letter(index + 1),
                employee_header=str(header).strip(),
                reason='Active employee not found for this column header'
            )

    if date_column_index is None or not employee_columns:
        return False

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if not row or not any(row):
            continue
        try:
            entry_date = _parse_excel_date(row[date_column_index])
        except ValueError:
            # Ignore report footer rows such as Billable Efforts or Accrual rows.
            continue

        # If a start/end date range is provided, skip any rows outside it.
        if start_date and end_date:
            if entry_date < start_date or entry_date > end_date:
                continue

        for column_index, employee, employee_header in employee_columns:
            raw_hours = row[column_index] if column_index < len(row) else None
            column_letter = get_column_letter(column_index + 1)
            try:
                hours = _parse_hours(raw_hours)
                _save_client_timesheet_entry(
                    employee=employee,
                    entry_date=entry_date,
                    hours=hours,
                    overwrite_drafts=overwrite_drafts,
                    results=results,
                    comments=f'Imported from client timesheet ({employee_header})',
                    row_idx=row_idx,
                    column_letter=column_letter,
                    employee_header=employee_header,
                    raw_hours=raw_hours
                )
            except Exception as exc:
                results['errors'].append(f'Row {row_idx}, {employee_header}: {str(exc)}')
                _record_client_import_issue(
                    results, 'failed', row_idx, column_letter, employee_header,
                    employee=employee, entry_date=entry_date, hours=raw_hours, reason=str(exc)
                )

    return True


def _import_client_timesheet_rows(worksheet, header_map, results, overwrite_drafts, data_start_row=2, start_date=None, end_date=None):
    """Compatibility import for employee/date/hours row-based sheets."""
    employee_lookup = _build_employee_lookup()
    employee_key = 'employee_id' if 'employee_id' in header_map else 'email'
    missing_headers = [
        header for header in [employee_key, 'date', 'hours']
        if header not in header_map
    ]
    if 'employee_id' not in header_map and 'email' not in header_map:
        missing_headers.insert(0, 'employee_id or email')

    if missing_headers:
        results['success'] = False
        results['message'] = 'Missing required columns in client timesheet file.'
        results['errors'].append(f"Missing columns: {', '.join(missing_headers)}")
        return

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if not row or not any(row):
            continue

        try:
            lookup_value = str(row[header_map[employee_key]] or '').strip()
            employee = _find_employee(lookup_value, employee_lookup)
            if not employee:
                raise ValueError(f'Active employee not found for {employee_key} "{lookup_value}".')

            entry_date = _parse_excel_date(row[header_map['date']])
            # Skip rows outside provided date range
            if start_date is not None and end_date is not None:
                if entry_date < start_date or entry_date > end_date:
                    continue
            raw_hours = row[header_map['hours']]
            hours = _parse_hours(raw_hours)
            comments = ''
            if 'comments' in header_map and row[header_map['comments']]:
                comments = str(row[header_map['comments']]).strip()
            elif 'comment' in header_map and row[header_map['comment']]:
                comments = str(row[header_map['comment']]).strip()

            _save_client_timesheet_entry(
                employee=employee,
                entry_date=entry_date,
                hours=hours,
                overwrite_drafts=overwrite_drafts,
                results=results,
                comments=comments,
                row_idx=row_idx,
                column_letter=get_column_letter(header_map['hours'] + 1),
                employee_header=lookup_value,
                raw_hours=raw_hours
            )
        except Exception as exc:
            results['errors'].append(f'Row {row_idx}: {str(exc)}')
            _record_client_import_issue(
                results, 'failed', row_idx, '', reason=str(exc)
            )


def import_client_timesheet_entries(file_obj, overwrite_drafts=True, start_date=None, end_date=None, notes=''):
    """
    Import client timesheet entries from an XLSX file.

    Expected client format:
    - Row 1 is ignored
    - Row 2 contains one Date column and employee columns
    - One column per employee, where the header is employee name, id, email, or
      "Name (Employee ID)"
    - Row 3 onward contains daily hours
    - Positive hour cells become TimesheetEntry rows using CLIENT_TIMESHEET_IMPORT_STATUS
    """
    import_logger.info('Starting client timesheet entry import')

    workbook = None
    try:
        workbook = load_workbook(file_obj, data_only=True)
        worksheet = workbook.active
        results = {
            'success': True,
            'message': 'Client timesheet import completed successfully',
            'created_count': 0,
            'updated_count': 0,
            'skipped_count': 0,
            'errors': [],
            'skipped_records': [],
            'failed_records': [],
            'exception_report_path': '',
            'import_status': _get_client_timesheet_import_status(),
            'import_start_date': start_date,
            'import_end_date': end_date,
            'compoff_created_count': 0,
            'compoff_deleted_count': 0,
            'compoff_taken_count': 0,
            'accrual_created_count': 0,
            '_affected_employee_dates': set(),
        }

        header_row = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
        header_row_number = 2
        data_start_row = 3
        if not header_row or not any(header_row):
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            header_row_number = 1
            data_start_row = 2

        if not header_row:
            return {
                'success': False,
                'message': 'The uploaded file is empty.',
                'created_count': 0,
                'updated_count': 0,
                'skipped_count': 0,
                'errors': ['Missing header row.'],
            }

        header_map = {
            _normalize_header(column): index
            for index, column in enumerate(header_row)
            if column is not None
        }

        is_row_based_format = (
            'date' in header_map
            and 'hours' in header_map
            and ('employee_id' in header_map or 'email' in header_map)
        )

        if is_row_based_format:
            _import_client_timesheet_rows(
                worksheet, header_map, results, overwrite_drafts, data_start_row=data_start_row,
                start_date=start_date, end_date=end_date
            )
        elif 'date' in header_map:
            imported = _import_client_timesheet_matrix(
                worksheet, header_row, results, overwrite_drafts,
                data_start_row=data_start_row,
                header_row_number=header_row_number,
                start_date=start_date, end_date=end_date
            )
            if not imported:
                _import_client_timesheet_rows(
                    worksheet, header_map, results, overwrite_drafts, data_start_row=data_start_row,
                    start_date=start_date, end_date=end_date
                )
        else:
            fallback_header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            fallback_header_map = {
                _normalize_header(column): index
                for index, column in enumerate(fallback_header_row or [])
                if column is not None
            }
            fallback_is_row_based_format = (
                'date' in fallback_header_map
                and 'hours' in fallback_header_map
                and ('employee_id' in fallback_header_map or 'email' in fallback_header_map)
            )
            if fallback_is_row_based_format:
                _import_client_timesheet_rows(
                    worksheet, fallback_header_map, results, overwrite_drafts, data_start_row=2,
                    start_date=start_date, end_date=end_date
                )
            elif 'date' in fallback_header_map:
                imported = _import_client_timesheet_matrix(
                    worksheet, fallback_header_row or [], results, overwrite_drafts,
                    data_start_row=2,
                    header_row_number=1,
                    start_date=start_date, end_date=end_date
                )
                if not imported:
                    _import_client_timesheet_rows(
                        worksheet, fallback_header_map, results, overwrite_drafts, data_start_row=2,
                        start_date=start_date, end_date=end_date
                    )
            else:
                _import_client_timesheet_rows(
                    worksheet, header_map, results, overwrite_drafts, data_start_row=data_start_row,
                    start_date=start_date, end_date=end_date
                )

        _auto_apply_pending_compoffs_for_import(results)
        _auto_adjust_pending_accruals_for_import(results)

        if results['errors']:
            results['success'] = False
            results['message'] = f"Client timesheet import completed with {len(results['errors'])} errors"
            import_logger.warning(results['message'])
        else:
            import_logger.info(
                f"Client timesheet import completed. Created={results['created_count']}, "
                f"Updated={results['updated_count']}, Skipped={results['skipped_count']}, "
                f"CompOffsCreated={results['compoff_created_count']}, "
                f"CompOffsDeleted={results['compoff_deleted_count']}, "
                f"CompOffsTaken={results['compoff_taken_count']}, "
                f"AccrualsCreated={results.get('accrual_created_count', 0)}"
            )

        results['exception_report_path'] = _write_client_import_exception_report(results)
        results.pop('_affected_employee_dates', None)
        return results
    except Exception as exc:
        import_logger.error(f"Failed to import client timesheet file: {str(exc)}", exc_info=True)
        return {
            'success': False,
            'message': f'Failed to import client timesheet file: {str(exc)}',
            'created_count': 0,
            'updated_count': 0,
            'skipped_count': 0,
            'errors': [str(exc)],
            'skipped_records': [],
            'failed_records': [{
                'type': 'FAILED',
                'row': '',
                'column': '',
                'employee': '',
                'employee_header': '',
                'date': '',
                'hours': '',
                'reason': str(exc),
            }],
            'exception_report_path': '',
            'import_status': getattr(settings, 'CLIENT_TIMESHEET_IMPORT_STATUS', 'DRAFT'),
            'compoff_created_count': 0,
            'compoff_deleted_count': 0,
            'compoff_taken_count': 0,
        }
    finally:
        if workbook:
            workbook.close()


def import_holiday_file(file_obj):
    """
    Import holiday data from an Excel file.

    Expected columns:
    - S. No.
    - Date
    - Day
    - Holiday
    - Holiday Type
    - Applicability

    Column mapping:
    - Holiday -> Holiday.name
    - Holiday Type -> Holiday.holiday_type
    - Applicability -> Holiday.location
    """
    holiday_type_map = {
        'public holiday': 'PUBLIC_HOLIDAY',
        'us holiday': 'US_HOLIDAY',
        'special holiday': 'SPECIAL_HOLIDAY',
    }

    try:
        workbook = load_workbook(file_obj)
        worksheet = workbook.active
        results = {
            'success': True,
            'message': 'Holiday import completed successfully',
            'created_count': 0,
            'updated_count': 0,
            'errors': [],
        }

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return {
                'success': False,
                'message': 'The uploaded file is empty.',
                'errors': ['Missing header row.'],
            }

        header_map = {
            str(column).strip().lower(): index
            for index, column in enumerate(header_row)
            if column is not None
        }
        required_headers = ['date', 'holiday', 'holiday type', 'applicability']
        missing_headers = [header for header in required_headers if header not in header_map]
        if missing_headers:
            return {
                'success': False,
                'message': 'Missing required columns in holiday import file.',
                'errors': [f"Missing columns: {', '.join(missing_headers)}"],
            }

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[header_map['date']] or not row[header_map['holiday']]:
                continue

            try:
                date_value = row[header_map['date']]
                if isinstance(date_value, datetime):
                    holiday_date = date_value.date()
                elif hasattr(date_value, 'year') and hasattr(date_value, 'month') and hasattr(date_value, 'day'):
                    holiday_date = date_value
                else:
                    holiday_date = datetime.strptime(str(date_value).strip(), '%Y-%m-%d').date()

                holiday_name = str(row[header_map['holiday']]).strip()
                holiday_type_label = str(row[header_map['holiday type']] or '').strip().lower()
                holiday_type = holiday_type_map.get(holiday_type_label)
                if not holiday_type:
                    raise ValueError(
                        'Holiday Type must be "Public Holiday", "US Holiday", or "Special Holiday".'
                    )
                location = str(row[header_map['applicability']]).strip() if row[header_map['applicability']] else ''

                holiday, created = Holiday.objects.update_or_create(
                    date=holiday_date,
                    name=holiday_name,
                    location=location,
                    defaults={
                        'holiday_type': holiday_type,
                    }
                )

                if created:
                    results['created_count'] += 1
                else:
                    results['updated_count'] += 1
            except Exception as exc:
                results['errors'].append(f'Row {row_idx}: {str(exc)}')

        workbook.close()

        if results['errors']:
            results['success'] = False
            results['message'] = f"Holiday import completed with {len(results['errors'])} errors"

        return results
    except Exception as exc:
        import_logger.error(f"Failed to import holiday file: {str(exc)}", exc_info=True)
        return {
            'success': False,
            'message': f'Failed to import holiday file: {str(exc)}',
            'errors': [str(exc)],
        }


def import_timesheet_file(file_path, import_date=None):
    """
    Import timesheet data from XLSX file using bulk_create for performance optimization.
    Expected sheets: Timesheet Summary, Timesheet Details, Attendance Summary, Attendance Details
    
    Uses bulk_create() for batch insert operations to reduce database I/O calls significantly.
    Batch size of 1000 records per insert to balance memory usage and performance.
    
    Args:
        file_path: Path to the XLSX file
        import_date: Date/Month of the import (optional, for filtering later)
    """
    import_logger.info(f"Starting timesheet import from file: {file_path}, import_date: {import_date}")
    
    try:
        workbook = load_workbook(file_path)
        results = {
            'success': True,
            'message': 'Import completed successfully',
            'summary_count': 0,
            'details_count': 0,
            'attendance_summary_count': 0,
            'attendance_details_count': 0,
            'errors': []
        }
        
        import_logger.info(f"Excel file loaded. Sheets found: {workbook.sheetnames}")
        BATCH_SIZE = 1000  # Process records in batches of 1000

        # Import Timesheet Summary
        if 'Timesheet Summary' in workbook.sheetnames:
            try:
                ws = workbook['Timesheet Summary']
                summaries_to_create = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Skip header
                        continue
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    try:
                        summary = TimesheetSummary(
                            email_id=str(row[0]).strip(),
                            team_member=str(row[1]).strip(),
                            project_id=int(row[2]),
                            project=str(row[3]).strip(),
                            total_hours=Decimal(str(row[4])),
                            import_date=import_date
                        )
                        summaries_to_create.append(summary)
                    except Exception as e:
                        results['errors'].append(f"Timesheet Summary Row {row_idx}: {str(e)}")
                
                # Bulk insert in batches
                if summaries_to_create:
                    TimesheetSummary.objects.bulk_create(summaries_to_create, batch_size=BATCH_SIZE)
                    results['summary_count'] = len(summaries_to_create)
                    import_logger.info(f"Bulk created {results['summary_count']} Timesheet Summary records")
                    
            except Exception as e:
                results['errors'].append(f"Error processing Timesheet Summary: {str(e)}")

        # Import Timesheet Details
        if 'Timesheet Details' in workbook.sheetnames:
            try:
                ws = workbook['Timesheet Details']
                details_to_create = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Skip header
                        continue
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    try:
                        # Parse date
                        date_val = row[4]
                        if isinstance(date_val, datetime):
                            date_val = date_val.date()
                        else:
                            date_val = datetime.strptime(str(date_val), '%Y-%m-%d').date()
                        
                        detail = TimesheetDetails(
                            email_id=str(row[0]).strip(),
                            team_member=str(row[1]).strip(),
                            project_id=int(row[2]),
                            project=str(row[3]).strip(),
                            date=date_val,
                            time_logged=Decimal(str(row[5])),
                            comment=str(row[6]).strip() if row[6] else '',
                            import_date=import_date
                        )
                        details_to_create.append(detail)
                    except Exception as e:
                        results['errors'].append(f"Timesheet Details Row {row_idx}: {str(e)}")
                
                # Bulk insert in batches
                if details_to_create:
                    TimesheetDetails.objects.bulk_create(details_to_create, batch_size=BATCH_SIZE)
                    results['details_count'] = len(details_to_create)
                    import_logger.info(f"Bulk created {results['details_count']} Timesheet Details records")
                    
            except Exception as e:
                results['errors'].append(f"Error processing Timesheet Details: {str(e)}")

        # Import Attendance Summary
        if 'Attendance Summary' in workbook.sheetnames:
            try:
                ws = workbook['Attendance Summary']
                att_summaries_to_create = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Skip header
                        continue
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    try:
                        summary = AttendanceSummary(
                            email_id=str(row[0]).strip(),
                            team_member=str(row[1]).strip(),
                            business_hours=Decimal(str(row[2])),
                            available_hours=Decimal(str(row[3])),
                            project_hours=Decimal(str(row[4])),
                            import_date=import_date
                        )
                        att_summaries_to_create.append(summary)
                    except Exception as e:
                        results['errors'].append(f"Attendance Summary Row {row_idx}: {str(e)}")
                
                # Bulk insert in batches
                if att_summaries_to_create:
                    AttendanceSummary.objects.bulk_create(att_summaries_to_create, batch_size=BATCH_SIZE)
                    results['attendance_summary_count'] = len(att_summaries_to_create)
                    import_logger.info(f"Bulk created {results['attendance_summary_count']} Attendance Summary records")
                    
            except Exception as e:
                results['errors'].append(f"Error processing Attendance Summary: {str(e)}")

        # Import Attendance Details
        if 'Attendance Details' in workbook.sheetnames:
            try:
                ws = workbook['Attendance Details']
                att_details_to_create = []
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Skip header
                        continue
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    try:
                        # Parse date
                        date_val = row[2]
                        if isinstance(date_val, datetime):
                            date_val = date_val.date()
                        else:
                            date_val = datetime.strptime(str(date_val), '%Y-%m-%d').date()
                        
                        detail = AttendanceDetails(
                            email_id=str(row[0]).strip(),
                            team_member=str(row[1]).strip(),
                            date=date_val,
                            business_hours=Decimal(str(row[3])),
                            available_hours=Decimal(str(row[4])),
                            remarks=str(row[5]).strip() if row[5] else '',
                            import_date=import_date
                        )
                        att_details_to_create.append(detail)
                    except Exception as e:
                        results['errors'].append(f"Attendance Details Row {row_idx}: {str(e)}")
                
                # Bulk insert in batches
                if att_details_to_create:
                    AttendanceDetails.objects.bulk_create(att_details_to_create, batch_size=BATCH_SIZE)
                    results['attendance_details_count'] = len(att_details_to_create)
                    import_logger.info(f"Bulk created {results['attendance_details_count']} Attendance Details records")
                    
            except Exception as e:
                results['errors'].append(f"Error processing Attendance Details: {str(e)}")

        if results['errors']:
            results['success'] = False
            results['message'] = f"Import completed with {len(results['errors'])} errors"
            import_logger.warning(f"Import completed with {len(results['errors'])} errors: {results['errors']}")
        else:
            results['message'] = 'Import completed successfully'
            import_logger.info(f"Import completed successfully (using bulk_create). Summary={results['summary_count']}, Details={results['details_count']}, "
                              f"AttSummary={results['attendance_summary_count']}, AttDetails={results['attendance_details_count']}")
        
        workbook.close()
        return results

    except Exception as e:
        import_logger.error(f"Failed to import file: {str(e)}", exc_info=True)
        return {
            'success': False,
            'message': f"Failed to import file: {str(e)}",
            'errors': [str(e)]
        }

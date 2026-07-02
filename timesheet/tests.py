from datetime import date, timedelta
from io import BytesIO
import glob
import os
from unittest.mock import patch

from django.contrib.auth.models import Group, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.conf import settings
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook

from timesheet.forms import TimesheetEntryForm, TimesheetImportForm
from timesheet.utils import import_client_timesheet_entries, import_employee_file, import_holiday_file

from timesheet.models import Accrual, CompOff, Employee, Holiday, Location, Project, TimesheetEntry, TimesheetImportLog


def get_location(name):
    return Location.objects.get_or_create(name=name)[0]


class ProjectModelTest(TestCase):
    def setUp(self):
        self.project = Project.objects.create(
            project_id=101,
            department_name='Risk Tech',
            project='SMBC1',
            project_code=1000,
            manager='Ranzith',
        )

    def test_project_creation(self):
        self.assertIsInstance(self.project, Project)
        self.assertEqual(self.project.department_name, 'Risk Tech')
        self.assertEqual(self.project.project, 'SMBC1')
        self.assertEqual(str(self.project), 'SMBC1 - Risk Tech (ID: 101)')


class EmployeeModelTest(TestCase):
    def setUp(self):
        self.project = Project.objects.create(
            project_id=201,
            department_name='Risk Tech',
            project='SMBC1',
            project_code=1001,
            manager='Ranzith',
        )
        self.employee = Employee.objects.create(
            name='John Doe',
            employee_id='001',
            email='john@example.com',
            project=self.project,
            location=get_location('Indore'),
        )

    def test_employee_creation(self):
        self.assertIsInstance(self.employee, Employee)
        self.assertEqual(self.employee.project, self.project)
        self.assertEqual(str(self.employee), 'John Doe (001)')

    def test_employee_unique_id(self):
        with self.assertRaises(Exception):
            Employee.objects.create(
                name='Jane Doe',
                employee_id='001',
                email='jane@example.com',
                project=self.project,
            )


class EmployeeImportTests(TestCase):
    def _build_employee_workbook_file(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['name', 'employee_id', 'email', 'project', 'location', 'is_active'])
        for row in rows:
            worksheet.append(row)

        file_obj = BytesIO()
        workbook.save(file_obj)
        workbook.close()
        file_obj.seek(0)
        return file_obj

    def test_import_employee_file_creates_project_employee_and_user(self):
        file_obj = self._build_employee_workbook_file([
            ['Priya Sharma', 'E1001', 'priya.sharma@test.com', 'CHIP', 'Indore', 'TRUE'],
        ])

        result = import_employee_file(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['created_count'], 1)
        self.assertEqual(result['user_created_count'], 1)
        self.assertEqual(result['project_created_count'], 1)

        employee = Employee.objects.select_related('project', 'user').get(employee_id='E1001')
        self.assertEqual(employee.name, 'Priya Sharma')
        self.assertEqual(employee.project.project, 'CHIP')
        self.assertEqual(employee.project.manager, 'CHIP')
        self.assertEqual(employee.user.username, 'priya.sharma')
        self.assertEqual(employee.user.email, 'priya.sharma@test.com')
        self.assertTrue(employee.user.groups.filter(name='Employee').exists())

    def test_import_employee_file_updates_existing_employee_and_user(self):
        project = Project.objects.create(
            project_id=1001,
            department_name='Risk Tech',
            project='Old Project',
            project_code=1000,
            manager='Old Project',
        )
        user = User.objects.create_user(
            username='amit.kumar',
            email='old@test.com',
            password='amit.kumar',
        )
        employee = Employee.objects.create(
            user=user,
            name='Amit Kumar',
            employee_id='E1002',
            email='old@test.com',
            project=project,
            location=get_location('Pune'),
            is_active=True,
        )
        file_obj = self._build_employee_workbook_file([
            ['Amit Kumar', 'E1002', 'amit.kumar@test.com', 'ESG', 'Mumbai', 'NO'],
        ])

        result = import_employee_file(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['created_count'], 0)
        self.assertEqual(result['updated_count'], 1)
        employee.refresh_from_db()
        user.refresh_from_db()
        self.assertEqual(employee.email, 'amit.kumar@test.com')
        self.assertEqual(employee.location.name, 'Mumbai')
        self.assertFalse(employee.is_active)
        self.assertEqual(employee.project.project, 'ESG')
        self.assertEqual(user.email, 'amit.kumar@test.com')
        self.assertFalse(user.is_active)

    def test_employee_admin_import_accepts_xlsx_upload(self):
        admin_user = User.objects.create_superuser(
            username='admin_employee_import',
            password='testpass123',
            email='admin@test.com',
        )
        file_obj = self._build_employee_workbook_file([
            ['Neha Patel', 'E1003', 'neha.patel@test.com', 'SDET', 'Indore', 'YES'],
        ])
        upload = SimpleUploadedFile(
            'employees.xlsx',
            file_obj.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        self.client.login(username='admin_employee_import', password='testpass123')
        response = self.client.post(
            reverse('admin:timesheet_employee_import'),
            {'file': upload},
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Employee.objects.filter(employee_id='E1003', user__username='neha.patel').exists())


class AccrualPageAccessTests(TestCase):
    def setUp(self):
        self.project = Project.objects.create(
            project_id=501,
            department_name='Risk Tech',
            project='SMBC1',
            project_code=5001,
            manager='Ranzith',
        )
        self.employee_group, _ = Group.objects.get_or_create(name='Employee')
        self.user = User.objects.create_user(
            username='employee_accrual_user',
            email='employee_accrual@test.com',
            password='testpass123',
        )
        self.user.groups.add(self.employee_group)
        self.employee = Employee.objects.create(
            user=self.user,
            name='Accrual Employee',
            employee_id='E5001',
            email='employee_accrual@test.com',
            project=self.project,
            location=get_location('Indore'),
        )
        self.accrual = Accrual.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 5),
            adjusted_date=date(2026, 4, 7),
            adjustment_reason='LEAVE_TAKEN',
            status='PENDING',
            notes='Test accrual',
        )

    def test_employee_can_view_their_own_accruals(self):
        self.client.login(username='employee_accrual_user', password='testpass123')

        response = self.client.get(reverse('accrual_list'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Accruals')
        self.assertEqual(list(response.context['accruals']), [self.accrual])

    def test_admin_can_open_accrual_add_form(self):
        admin_user = User.objects.create_superuser(
            username='admin_accrual_form',
            password='testpass123',
            email='admin_accrual_form@test.com',
        )
        self.client.login(username='admin_accrual_form', password='testpass123')

        response = self.client.get(reverse('accrual_add'))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'timesheet/accrual_form.html')


class AccrualGenerationCommandTests(TestCase):
    def setUp(self):
        self.project = Project.objects.create(
            project_id=701,
            department_name='Risk Tech',
            project='SMBC1',
            project_code=7001,
            manager='Ranzith',
        )
        self.employee = Employee.objects.create(
            name='Accrual Migrator',
            employee_id='E7001',
            email='accrual.migrator@test.com',
            project=self.project,
            location=get_location('Indore'),
        )

    def test_generate_accruals_from_existing_compoffs(self):
        CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 5),
            compoff_date=date(2026, 4, 7),
            status='TAKEN',
            notes='Existing CompOff',
        )
        CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 12),
            compoff_date=None,
            status='PENDING',
            notes='Pending CompOff',
        )

        call_command('generate_accruals_from_compoffs', verbosity=0)

        accruals = Accrual.objects.filter(employee=self.employee).order_by('working_date')

        self.assertEqual(accruals.count(), 2)
        first = accruals[0]
        second = accruals[1]

        self.assertEqual(first.working_date, date(2026, 4, 5))
        self.assertEqual(first.adjusted_date, date(2026, 4, 7))
        self.assertEqual(first.status, 'ADJUSTED')
        self.assertIn('Existing CompOff', first.notes)

        self.assertEqual(second.working_date, date(2026, 4, 12))
        self.assertIsNone(second.adjusted_date)
        self.assertEqual(second.status, 'PENDING')


class CompOffModelTest(TestCase):
    def setUp(self):
        self.project = Project.objects.create(
            project_id=301,
            department_name='Risk Tech',
            project='SMBC1',
            project_code=1002,
            manager='Ranzith',
        )
        self.employee = Employee.objects.create(
            name='John Doe',
            employee_id='002',
            email='john@example.com',
            project=self.project,
            location=get_location('Indore'),
        )

    def test_compoff_defaults_to_pending_without_compoff_date(self):
        compoff = CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 5),
            notes='Weekend work',
        )

        self.assertEqual(compoff.status, 'PENDING')
        self.assertIsNone(compoff.compoff_date)

    def test_compoff_becomes_taken_for_past_compoff_date(self):
        compoff = CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 5),
            compoff_date=date.today() - timedelta(days=1),
            notes='Used compoff',
        )

        self.assertEqual(compoff.status, 'TAKEN')


class TimesheetEntryModelTest(TestCase):
    def setUp(self):
        self.project = Project.objects.create(
            project_id=401,
            department_name='Risk Tech',
            project='SMBC1',
            project_code=1003,
            manager='Ranzith',
        )
        self.employee = Employee.objects.create(
            name='Alice',
            employee_id='003',
            email='alice@example.com',
            project=self.project,
            location=get_location('Indore'),
        )

    def test_timesheet_entry_creation(self):
        entry = TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 1),
            project='Client Project',
            hours=8,
            comments='Worked on feature',
            status='DRAFT',
        )

        self.assertIsInstance(entry, TimesheetEntry)
        self.assertEqual(entry.employee.name, 'Alice')
        self.assertEqual(entry.status, 'DRAFT')


class HolidayModelTest(TestCase):
    def test_holiday_creation(self):
        holiday = Holiday.objects.create(
            name='Indore Holiday',
            date=date(2026, 5, 1),
            holiday_type='PUBLIC_HOLIDAY',
            location=get_location('Indore'),
        )

        self.assertIsInstance(holiday, Holiday)
        self.assertEqual(
            str(holiday),
            'Indore Holiday - 2026-05-01 (PUBLIC_HOLIDAY)'
        )


class HolidayImportTests(TestCase):
    def _build_holiday_workbook_file(self, rows, headers=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(headers or ['S. No.', 'Date', 'Day', 'Holiday', 'Holiday Type', 'Applicability'])
        for row in rows:
            worksheet.append(row)

        file_obj = BytesIO()
        workbook.save(file_obj)
        workbook.close()
        file_obj.seek(0)
        return file_obj

    def test_import_holiday_file_stores_uploaded_holiday_types(self):
        file_obj = self._build_holiday_workbook_file([
            [1, '2026-07-04', 'Saturday', 'Independence Day', 'US Holiday', 'Indore'],
            [2, '2026-08-15', 'Saturday', 'Independence Day India', 'Public Holiday', 'Indore'],
        ])

        result = import_holiday_file(file_obj)

        self.assertTrue(result['success'], result['errors'])
        self.assertEqual(result['created_count'], 2)
        self.assertEqual(
            Holiday.objects.get(name='Independence Day').holiday_type,
            'US_HOLIDAY'
        )
        self.assertEqual(
            Holiday.objects.get(name='Independence Day India').holiday_type,
            'PUBLIC_HOLIDAY'
        )

    def test_import_holiday_file_requires_holiday_type_column(self):
        file_obj = self._build_holiday_workbook_file(
            [[1, '2026-07-04', 'Saturday', 'Independence Day', 'Indore']],
            headers=['S. No.', 'Date', 'Day', 'Holiday', 'Applicability']
        )

        result = import_holiday_file(file_obj)

        self.assertFalse(result['success'])
        self.assertIn('holiday type', result['errors'][0])

    def test_import_holiday_file_rejects_unknown_holiday_type(self):
        file_obj = self._build_holiday_workbook_file([
            [1, '2026-07-04', 'Saturday', 'Independence Day', 'Bank Holiday', 'Indore'],
        ])

        result = import_holiday_file(file_obj)

        self.assertFalse(result['success'])
        self.assertIn('Public Holiday', result['errors'][0])
        self.assertIn('US Holiday', result['errors'][0])


class EmployeeProjectSelectionTests(TestCase):
    def setUp(self):
        self.employee_group, _ = Group.objects.get_or_create(name='Employee')
        self.admin_user = User.objects.create_user(
            username='admin_timesheet',
            password='testpass123',
            is_staff=True,
        )
        self.user = User.objects.create_user(username='bob', password='testpass123')
        self.user.groups.add(self.employee_group)
        self.project = Project.objects.create(
            project_id=601,
            department_name='Risk Tech',
            project='Risk Tech',
            project_code=1601,
            manager='Manager Two',
        )
        self.employee = Employee.objects.create(
            user=self.user,
            name='Bob',
            employee_id='E601',
            email='bob@example.com',
            project=self.project,
            location=get_location('Indore'),
        )

    def test_timesheet_entry_form_limits_project_to_employee_assignment(self):
        form = TimesheetEntryForm(employee=self.employee)

        self.assertEqual(
            list(form.fields['project'].choices),
            [('Risk Tech', 'Risk Tech')]
        )

    def test_generate_timesheet_rejects_project_outside_employee_assignment(self):
        self.client.login(username='bob', password='testpass123')

        response = self.client.post(
            reverse('generate_timesheet_weekdays'),
            {
                'month': '2026-04',
                'from_date': '2026-04-01',
                'to_date': '2026-04-02',
                'project': 'Other Project',
                'hours_per_day': '8',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(
            TimesheetEntry.objects.filter(
                employee=self.employee,
                date=date(2026, 4, 1),
            ).exists()
        )

    def test_generate_timesheet_creates_and_highlights_us_holiday_entry(self):
        Holiday.objects.create(
            name='US Holiday',
            date=date(2026, 4, 3),
            holiday_type='US_HOLIDAY',
            location='Indore',
        )
        self.client.login(username='bob', password='testpass123')

        response = self.client.post(
            reverse('generate_timesheet_weekdays'),
            {
                'month': '2026-04',
                'from_date': '2026-04-03',
                'to_date': '2026-04-03',
                'project': 'Risk Tech',
                'hours_per_day': '8',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            TimesheetEntry.objects.filter(
                employee=self.employee,
                date=date(2026, 4, 3),
                project='Risk Tech',
                hours=8,
                status='DRAFT',
            ).exists()
        )

        response = self.client.get(reverse('timesheet_entry_list'), {'month': '2026-04'})
        self.assertIn(date(2026, 4, 3), response.context['holiday_dates'])

    def test_admin_can_edit_submitted_timesheet_and_sync_pending_compoff(self):
        entry = TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 4),
            project='Risk Tech',
            hours=8,
            comments='Imported from client timesheet',
            status='SUBMITTED',
        )
        CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 4),
            status='PENDING',
            notes='Auto-generated Comp-Off from client timesheet import',
        )

        self.client.login(username='admin_timesheet', password='testpass123')
        response = self.client.post(
            reverse('timesheet_entry_edit', args=[entry.pk]) + f'?month=2026-04&employee={self.employee.pk}',
            {
                'date': '2026-04-06',
                'project': 'Risk Tech',
                'hours': '7',
                'comments': 'Admin correction',
            },
        )

        self.assertEqual(response.status_code, 302)
        entry.refresh_from_db()
        self.assertEqual(entry.date, date(2026, 4, 6))
        self.assertEqual(entry.hours, 7)
        self.assertEqual(entry.status, 'SUBMITTED')
        self.assertFalse(
            CompOff.objects.filter(
                employee=self.employee,
                working_date=date(2026, 4, 4),
                status='PENDING',
            ).exists()
        )

    def test_admin_edit_submitted_timesheet_to_zero_marks_pending_compoff_taken(self):
        entry = TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 6),
            project='Risk Tech',
            hours=8,
            comments='Submitted hours',
            status='SUBMITTED',
        )
        compoff = CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 4),
            status='PENDING',
            notes='Weekend work',
        )

        self.client.login(username='admin_timesheet', password='testpass123')
        response = self.client.post(
            reverse('timesheet_entry_edit', args=[entry.pk]) + f'?month=2026-04&employee={self.employee.pk}',
            {
                'date': '2026-04-06',
                'project': 'Risk Tech',
                'hours': '0',
                'comments': 'Admin correction',
            },
        )

        self.assertEqual(response.status_code, 302)
        compoff.refresh_from_db()
        self.assertEqual(compoff.status, 'TAKEN')
        self.assertEqual(compoff.compoff_date, date(2026, 4, 6))

    def test_admin_can_delete_submitted_timesheet_and_pending_compoff(self):
        entry = TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 4),
            project='Risk Tech',
            hours=8,
            comments='Imported from client timesheet',
            status='SUBMITTED',
        )
        CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 4),
            status='PENDING',
            notes='Auto-generated Comp-Off from client timesheet import',
        )

        self.client.login(username='admin_timesheet', password='testpass123')
        response = self.client.delete(reverse('timesheet_entry_delete', args=[entry.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertFalse(TimesheetEntry.objects.filter(pk=entry.pk).exists())
        self.assertFalse(
            CompOff.objects.filter(
                employee=self.employee,
                working_date=date(2026, 4, 4),
                status='PENDING',
            ).exists()
        )

    def test_admin_delete_submitted_weekday_timesheet_marks_pending_compoff_taken(self):
        entry = TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 6),
            project='Risk Tech',
            hours=8,
            status='SUBMITTED',
        )
        compoff = CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 4),
            status='PENDING',
            notes='Weekend work',
        )

        self.client.login(username='admin_timesheet', password='testpass123')
        response = self.client.delete(reverse('timesheet_entry_delete', args=[entry.pk]))

        self.assertEqual(response.status_code, 200)
        compoff.refresh_from_db()
        self.assertEqual(compoff.status, 'TAKEN')
        self.assertEqual(compoff.compoff_date, date(2026, 4, 6))


    def test_employee_cannot_edit_submitted_timesheet(self):
        entry = TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 1),
            project='Risk Tech',
            hours=8,
            status='SUBMITTED',
        )

        self.client.login(username='bob', password='testpass123')
        response = self.client.get(reverse('timesheet_entry_edit', args=[entry.pk]))

        self.assertEqual(response.status_code, 302)


class TimesheetImportFormTests(TestCase):
    def test_import_month_accepts_year_month_value(self):
        upload = SimpleUploadedFile(
            'timesheet.xlsx',
            b'dummy',
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        form = TimesheetImportForm(
            data={'import_date': '2026-04', 'notes': ''},
            files={'file': upload}
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['import_date'], date(2026, 4, 1))


class ClientReportingRulesTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username='admin',
            password='testpass123',
            is_staff=True,
        )
        self.project = Project.objects.create(
            project_id=701,
            department_name='Risk Tech',
            project='Risk Tech',
            project_code=1701,
            manager='Manager Three',
        )
        self.employee = Employee.objects.create(
            name='Charlie',
            employee_id='E701',
            email='charlie@example.com',
            project=self.project,
            location=get_location('Indore'),
        )

    def test_client_reporting_hides_weekend_worked_hours(self):
        TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 4),
            project='Risk Tech',
            hours=8,
            status='SUBMITTED',
        )

        self.client.login(username='admin', password='testpass123')
        response = self.client.get(
            reverse('client_reporting'),
            {
                'manager': 'Manager Three',
                'start_date': '2026-04-04',
                'end_date': '2026-04-04',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['report_rows'][0]['hours'][0], '')
        self.assertEqual(response.context['employee_totals'][0], 0)

    def test_client_reporting_hides_us_holiday_worked_hours(self):
        Holiday.objects.create(
            name='US Holiday',
            date=date(2026, 4, 3),
            holiday_type='US_HOLIDAY',
            location='Indore',
        )
        TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 3),
            project='Risk Tech',
            hours=8,
            status='SUBMITTED',
        )

        self.client.login(username='admin', password='testpass123')
        response = self.client.get(
            reverse('client_reporting'),
            {
                'manager': 'Manager Three',
                'start_date': '2026-04-03',
                'end_date': '2026-04-03',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['report_rows'][0]['hours'][0], '')
        self.assertEqual(response.context['report_rows'][0]['date_type'], 'US Holiday')
        self.assertEqual(response.context['employee_totals'][0], 0)

    def test_client_reporting_shows_public_holiday_accrual_hours_for_adjusted_off_day(self):
        Holiday.objects.create(
            name='Public Holiday',
            date=date(2026, 4, 3),
            holiday_type='PUBLIC_HOLIDAY',
            location='Indore',
        )
        Accrual.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 3),
            adjusted_date=date(2026, 4, 3),
            adjustment_reason='PUBLIC_HOLIDAY_OFF',
            status='ADJUSTED',
            notes='Adjusted for public holiday off',
        )

        self.client.login(username='admin', password='testpass123')
        response = self.client.get(
            reverse('client_reporting'),
            {
                'manager': 'Manager Three',
                'start_date': '2026-04-03',
                'end_date': '2026-04-03',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['report_rows'][0]['hours'][0], 8.0)
        self.assertEqual(response.context['employee_totals'][0], 8.0)

    def test_client_reporting_footer_lists_accrual_dates(self):
        Accrual.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 8),
            status='PENDING',
            notes='Pending accrual',
        )
        Accrual.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 4),
            adjusted_date=date(2026, 4, 6),
            adjustment_reason='LEAVE_TAKEN',
            status='ADJUSTED',
            notes='Adjusted for leave taken',
        )
        Accrual.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 3),
            adjusted_date=date(2026, 4, 7),
            adjustment_reason='PUBLIC_HOLIDAY_OFF',
            status='ADJUSTED',
            notes='Adjusted for public holiday off',
        )

        self.client.login(username='admin', password='testpass123')
        response = self.client.get(
            reverse('client_reporting'),
            {
                'manager': 'Manager Three',
                'start_date': '2026-04-01',
                'end_date': '2026-04-30',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['accrual_pending_hours_list'][0], 8)
        self.assertEqual(response.context['accrual_days_list'][0], '08-Apr-2026')
        self.assertEqual(
            response.context['accrual_adjusted_list'][0],
            '06-Apr-2026, 07-Apr-2026'
        )

    def test_client_reporting_shows_taken_compoff_hours_on_adjusted_weekday(self):
        TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 5),
            project='Risk Tech',
            hours=10,
            status='SUBMITTED',
        )
        CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 5),
            compoff_date=date(2026, 4, 6),
            status='TAKEN',
        )

        self.client.login(username='admin', password='testpass123')
        response = self.client.get(
            reverse('client_reporting'),
            {
                'manager': 'Manager Three',
                'start_date': '2026-04-06',
                'end_date': '2026-04-06',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['report_rows'][0]['hours'][0], 10.0)
        self.assertTrue(response.context['report_rows'][0]['compoff_taken'][0])
        self.assertEqual(response.context['employee_totals'][0], 10.0)

    def test_client_reporting_email_uses_project_to_and_cc_recipients(self):
        self.project.to_email = 'client.to@example.com; sponsor@example.com'
        self.project.cc_email = 'client.cc@example.com; sponsor@example.com'
        self.project.save()
        TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 6),
            project='Risk Tech',
            hours=8,
            status='SUBMITTED',
        )

        self.client.login(username='admin', password='testpass123')
        with patch('timesheet.views.send_html_email_async') as mock_send:
            response = self.client.post(
                reverse('client_reporting'),
                {
                    'action': 'email_report',
                    'manager': 'Manager Three',
                    'start_date': '2026-04-06',
                    'end_date': '2026-04-06',
                },
            )

        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once()
        self.assertEqual(mock_send.call_args.args[3], ['client.to@example.com', 'sponsor@example.com'])
        self.assertEqual(mock_send.call_args.kwargs['cc'], ['client.cc@example.com'])

    def test_client_reporting_test_email_uses_entered_recipient_only(self):
        TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 6),
            project='Risk Tech',
            hours=8,
            status='SUBMITTED',
        )

        self.client.login(username='admin', password='testpass123')
        with patch('timesheet.views.send_html_email_async') as mock_send:
            response = self.client.post(
                reverse('client_reporting'),
                {
                    'action': 'test_email_report',
                    'manager': 'Manager Three',
                    'start_date': '2026-04-06',
                    'end_date': '2026-04-06',
                    'test_email': 'reviewer@example.com',
                },
            )

        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once()
        self.assertFalse(mock_send.call_args.args[0].startswith('[TEST] '))
        self.assertEqual(mock_send.call_args.args[3], ['reviewer@example.com'])
        self.assertEqual(mock_send.call_args.kwargs['cc'], [])


class AccrualSummaryFormattingTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username='admin_accrual',
            password='testpass123',
            is_staff=True,
        )
        self.project = Project.objects.create(
            project_id=801,
            department_name='Risk Tech',
            project='Risk Tech',
            project_code=1801,
            manager='Manager Four',
        )
        self.employee = Employee.objects.create(
            name='Dana',
            employee_id='E801',
            email='dana@example.com',
            project=self.project,
            location=get_location('Indore'),
        )

    def test_accrual_adjusted_message_includes_work_and_taken_dates(self):
        Accrual.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 5),
            adjusted_date=date(2026, 4, 7),
            adjustment_reason='LEAVE_TAKEN',
            status='ADJUSTED',
        )
        Accrual.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 12),
            adjusted_date=date(2026, 4, 14),
            adjustment_reason='PUBLIC_HOLIDAY_OFF',
            status='ADJUSTED',
        )

        self.client.login(username='admin_accrual', password='testpass123')
        response = self.client.get(
            reverse('accrual_summary'),
            {
                'start_date': '2026-04-01',
                'end_date': '2026-04-30',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context['accrual_data'][0]['adjusted_dates'],
            'Worked on 05-Apr-2026, 12-Apr-2026 adjusted against PTO on 07-Apr-2026, 14-Apr-2026'
        )

    def test_accrual_summary_shows_all_managers_in_separate_tables(self):
        second_project = Project.objects.create(
            project_id=802,
            department_name='Risk Tech',
            project='Risk QA',
            project_code=1802,
            manager='Manager Five',
        )
        Employee.objects.create(
            name='Quinn',
            employee_id='E802',
            email='quinn@example.com',
            role='QA',
            project=second_project,
            location=get_location('Indore'),
        )

        self.client.login(username='admin_accrual', password='testpass123')
        response = self.client.get(
            reverse('accrual_summary'),
            {
                'start_date': '2026-04-01',
                'end_date': '2026-04-30',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [table['manager'] for table in response.context['summary_tables']],
            ['Manager Five', 'Manager Four']
        )
        self.assertEqual(response.context['summary_tables'][0]['accrual_groups'][0]['role_label'], 'QA')
        self.assertEqual(response.context['summary_tables'][1]['accrual_groups'][0]['role_label'], 'DEV')

    @override_settings(
        ACCRUAL_SUMMARY_TO_EMAILS='accrual.to@example.com',
        ACCRUAL_SUMMARY_CC_EMAILS='accrual.cc@example.com',
    )
    def test_accrual_summary_email_uses_project_to_and_cc_recipients(self):
        self.client.login(username='admin_accrual', password='testpass123')
        with patch('timesheet.views.send_html_email_async') as mock_send:
            response = self.client.post(
                reverse('accrual_summary'),
                {
                    'action': 'email_report',
                    'start_date': '2026-04-01',
                    'end_date': '2026-04-30',
                },
            )

        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once()
        self.assertIn('April 30, 2026', mock_send.call_args.args[0])
        self.assertIn('April 30, 2026', mock_send.call_args.args[1])
        self.assertNotIn('2026-04-30', mock_send.call_args.args[1])
        self.assertEqual(mock_send.call_args.args[3], ['accrual.to@example.com'])
        self.assertEqual(mock_send.call_args.kwargs['cc'], ['accrual.cc@example.com'])

    def test_accrual_summary_test_email_uses_entered_recipient_only(self):
        self.client.login(username='admin_accrual', password='testpass123')
        with patch('timesheet.views.send_html_email_async') as mock_send:
            response = self.client.post(
                reverse('accrual_summary'),
                {
                    'action': 'test_email_report',
                    'start_date': '2026-04-01',
                    'end_date': '2026-04-30',
                    'test_email': 'reviewer@example.com',
                },
            )

        self.assertEqual(response.status_code, 200)
        mock_send.assert_called_once()
        self.assertFalse(mock_send.call_args.args[0].startswith('[TEST] '))
        self.assertEqual(mock_send.call_args.args[3], ['reviewer@example.com'])
        self.assertEqual(mock_send.call_args.kwargs['cc'], [])


class ClientTimesheetImportTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username='admin_import',
            password='testpass123',
            is_staff=True,
        )
        self.project = Project.objects.create(
            project_id=901,
            department_name='Risk Tech',
            project='Risk Tech',
            project_code=1901,
            manager='Manager Five',
        )
        self.employee = Employee.objects.create(
            name='Erin',
            employee_id='E901',
            email='erin@example.com',
            project=self.project,
            location=get_location('Indore'),
        )
        self.second_employee = Employee.objects.create(
            name='Sam',
            employee_id='E902',
            email='sam@example.com',
            project=self.project,
            location=get_location('Indore'),
        )

    def tearDown(self):
        for report_path in glob.glob(os.path.join(settings.BASE_DIR, 'import_reports', 'client_timesheet_import_exceptions_*.csv')):
            try:
                os.remove(report_path)
            except OSError:
                pass

    def _build_workbook_file(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['employee_id', 'date', 'project', 'hours', 'comments', 'status'])
        for row in rows:
            worksheet.append(row)

        file_obj = BytesIO()
        workbook.save(file_obj)
        workbook.close()
        file_obj.seek(0)
        return file_obj

    def _build_matrix_workbook_file(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Client Timesheet'])
        worksheet.append(['Date', 'Erin', 'Sam (E902)'])
        for row in rows:
            worksheet.append(row)

        file_obj = BytesIO()
        workbook.save(file_obj)
        workbook.close()
        file_obj.seek(0)
        return file_obj

    def test_import_client_timesheet_entries_creates_entries(self):
        file_obj = self._build_workbook_file([
            ['E901', '2026-04-01', 'Risk Tech', 8, 'Feature work', 'DRAFT'],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['created_count'], 1)
        entry = TimesheetEntry.objects.get(employee=self.employee)
        self.assertEqual(entry.date, date(2026, 4, 1))
        self.assertEqual(entry.project, 'Risk Tech')
        self.assertEqual(entry.hours, 8)
        self.assertEqual(entry.status, 'SUBMITTED')

    def test_import_client_timesheet_creates_compoff_for_weekend_work(self):
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-04', 8, 0],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['compoff_created_count'], 1)
        self.assertTrue(
            CompOff.objects.filter(
                employee=self.employee,
                working_date=date(2026, 4, 4),
                status='PENDING',
            ).exists()
        )

    def test_import_client_timesheet_creates_compoff_for_public_holiday_work(self):
        Holiday.objects.create(
            name='Holiday',
            date=date(2026, 4, 3),
            holiday_type='PUBLIC_HOLIDAY',
            location='Pune, Indore, Chennai',
        )
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-03', 8, 0],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['compoff_created_count'], 1)
        self.assertTrue(
            CompOff.objects.filter(
                employee=self.employee,
                working_date=date(2026, 4, 3),
                status='PENDING',
            ).exists()
        )

    def test_import_client_timesheet_creates_accrual_for_us_holiday_work(self):
        Holiday.objects.create(
            name='US Holiday',
            date=date(2026, 4, 3),
            holiday_type='US_HOLIDAY',
            location='Pune, Indore, Chennai',
        )
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-03', 8, 0],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['compoff_created_count'], 0)
        self.assertTrue(
            Accrual.objects.filter(
                employee=self.employee,
                working_date=date(2026, 4, 3),
                status='PENDING',
            ).exists()
        )
        self.assertFalse(
            CompOff.objects.filter(
                employee=self.employee,
                working_date=date(2026, 4, 3),
            ).exists()
        )

    def test_import_client_timesheet_matrix_imports_zero_and_blank_cells_as_zero_hours(self):
        file_obj = self._build_matrix_workbook_file([
            ['01-Apr-2026', 8, 0],
            ['02-Apr-2026', None, 6],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['created_count'], 4)
        self.assertEqual(result['skipped_count'], 0)
        self.assertEqual(len(result['skipped_records']), 0)
        self.assertFalse(result['exception_report_path'])
        self.assertTrue(
            TimesheetEntry.objects.filter(
                employee=self.employee,
                date=date(2026, 4, 1),
                hours=8,
                status='SUBMITTED',
            ).exists()
        )
        self.assertTrue(
            TimesheetEntry.objects.filter(
                employee=self.second_employee,
                date=date(2026, 4, 2),
                hours=6,
                status='SUBMITTED',
            ).exists()
        )
        self.assertTrue(
            TimesheetEntry.objects.filter(
                employee=self.second_employee,
                date=date(2026, 4, 1),
                hours=0,
                status='SUBMITTED',
            ).exists()
        )
        self.assertTrue(
            TimesheetEntry.objects.filter(
                employee=self.employee,
                date=date(2026, 4, 2),
                hours=0,
                status='SUBMITTED',
            ).exists()
        )

    def test_import_client_timesheet_overwrites_submitted_entry(self):
        TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 1),
            project='Risk Tech',
            hours=4,
            comments='Existing submitted',
            status='SUBMITTED',
        )
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-01', 8, 0],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['updated_count'], 1)
        entry = TimesheetEntry.objects.get(employee=self.employee, date=date(2026, 4, 1))
        self.assertEqual(entry.hours, 8)
        self.assertEqual(entry.status, 'SUBMITTED')

    def test_import_client_timesheet_blank_cell_clears_existing_hours_and_pending_compoff(self):
        TimesheetEntry.objects.create(
            employee=self.employee,
            date=date(2026, 4, 4),
            project='Risk Tech',
            hours=8,
            comments='Existing submitted',
            status='SUBMITTED',
        )
        CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 4),
            status='PENDING',
        )
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-04', None, 0],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['updated_count'], 1)
        self.assertEqual(result['compoff_deleted_count'], 1)
        entry = TimesheetEntry.objects.get(employee=self.employee, date=date(2026, 4, 4))
        self.assertEqual(entry.hours, 0)
        self.assertEqual(entry.status, 'SUBMITTED')
        self.assertFalse(
            CompOff.objects.filter(
                employee=self.employee,
                working_date=date(2026, 4, 4),
                status='PENDING',
            ).exists()
        )

    def test_import_client_timesheet_adjusts_pending_accrual_for_zero_weekday(self):
        Accrual.objects.create(
            employee=self.employee,
            working_date=date(2026, 3, 29),
            status='PENDING',
        )
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-06', 0, 8],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        accrual = Accrual.objects.get(employee=self.employee, working_date=date(2026, 3, 29))
        self.assertEqual(accrual.status, 'ADJUSTED')
        self.assertEqual(accrual.adjusted_date, date(2026, 4, 6))
        self.assertEqual(accrual.adjustment_reason, 'LEAVE_TAKEN')

    def test_import_client_timesheet_adjusts_pending_accrual_for_missing_weekdays_in_selected_period(self):
        Accrual.objects.create(
            employee=self.employee,
            working_date=date(2026, 3, 29),
            status='PENDING',
        )
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-03', 8, 0],
        ])

        result = import_client_timesheet_entries(
            file_obj,
            start_date=date(2026, 4, 1),
            end_date=date(2026, 4, 3),
        )

        self.assertTrue(result['success'])
        accrual = Accrual.objects.get(employee=self.employee, working_date=date(2026, 3, 29))
        self.assertEqual(accrual.status, 'ADJUSTED')
        self.assertEqual(accrual.adjusted_date, date(2026, 4, 1))
        self.assertEqual(accrual.adjustment_reason, 'LEAVE_TAKEN')

    def test_import_client_timesheet_zero_weekday_marks_pending_compoff_taken(self):
        compoff = CompOff.objects.create(
            employee=self.employee,
            working_date=date(2026, 4, 4),
            status='PENDING',
            notes='Weekend work',
        )
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-06', 0, 8],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['compoff_taken_count'], 1)
        compoff.refresh_from_db()
        self.assertEqual(compoff.status, 'TAKEN')
        self.assertEqual(compoff.compoff_date, date(2026, 4, 6))

    @override_settings(CLIENT_TIMESHEET_IMPORT_STATUS='DRAFT')
    def test_import_client_timesheet_status_can_be_configured(self):
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-04', 8, 0],
        ])

        result = import_client_timesheet_entries(file_obj)

        self.assertTrue(result['success'])
        self.assertEqual(result['import_status'], 'DRAFT')
        self.assertTrue(
            TimesheetEntry.objects.filter(
                employee=self.employee,
                date=date(2026, 4, 4),
                status='DRAFT',
            ).exists()
        )
        self.assertEqual(result['compoff_created_count'], 0)
        self.assertFalse(CompOff.objects.filter(employee=self.employee).exists())

    def test_import_client_timesheet_view_is_admin_only(self):
        user = User.objects.create_user(username='regular', password='testpass123')
        self.client.login(username='regular', password='testpass123')

        response = self.client.get(reverse('import_client_timesheet_entries'))

        self.assertEqual(response.status_code, 302)
        self.assertFalse(TimesheetEntry.objects.exists())

    def test_import_client_timesheet_reports_failed_and_skipped_records(self):
        file_obj = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Client Timesheet'])
        worksheet.append(['Date', 'Erin', 'Missing Employee'])
        worksheet.append(['2026-04-01', 14, 5])
        workbook.save(file_obj)
        workbook.close()
        file_obj.seek(0)

        result = import_client_timesheet_entries(file_obj)

        self.assertFalse(result['success'])
        self.assertEqual(len(result['failed_records']), 2)
        self.assertIn('Active employee not found', result['failed_records'][0]['reason'])
        self.assertIn('Hours must be between 0 and 12', result['failed_records'][1]['reason'])
        self.assertTrue(result['exception_report_path'])
        self.assertTrue(os.path.exists(os.path.join(settings.BASE_DIR, result['exception_report_path'])))

    def test_import_client_timesheet_view_accepts_xlsx_upload(self):
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-01', 8, 0],
        ])
        upload = SimpleUploadedFile(
            'client_timesheet.xlsx',
            file_obj.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        self.client.login(username='admin_import', password='testpass123')
        response = self.client.post(
            reverse('import_client_timesheet_entries'),
            {
                'file': upload,
                'import_date': '2026-04',
                'notes': 'April client import',
                'overwrite_drafts': 'on',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('result', response.context)
        self.assertTrue(TimesheetEntry.objects.filter(employee=self.employee).exists())

    def test_client_import_month_accepts_year_month_value(self):
        file_obj = self._build_matrix_workbook_file([
            ['2026-04-01', 8, 0],
        ])
        upload = SimpleUploadedFile(
            'client_timesheet.xlsx',
            file_obj.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        self.client.login(username='admin_import', password='testpass123')
        response = self.client.post(
            reverse('import_client_timesheet_entries'),
            {
                'file': upload,
                'import_date': '2026-04',
                'notes': 'April client import',
                'overwrite_drafts': 'on',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            TimesheetImportLog.objects.filter(import_date=date(2026, 4, 1)).exists()
        )
